"""
Convert enriched CVE CSV to a formatted Excel workbook with:
 - Color-coded columns: File/Path (azure), Function (light green), Subsystem (light orange)
 - AutoFilter on all columns
 - Data Validation dropdown on Subsystem and Severity columns
 - Frozen header row
 - Auto-fitted column widths
 - Optional SAST Report tab from Finite State .numbers export

Launches a simple GUI file browser to select:
 1. Enriched CVE CSV (required)
 2. SAST .numbers file (optional)
"""

import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Shared styles
# ---------------------------------------------------------------------------
AZURE_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
GREEN_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")

AZURE_HDR = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
GREEN_HDR = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
ORANGE_HDR = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
GREY_HDR = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
PURPLE_HDR = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

SEV_FILLS = {
    "critical": PatternFill(start_color="FF4D4D", end_color="FF4D4D", fill_type="solid"),
    "high": PatternFill(start_color="FF9933", end_color="FF9933", fill_type="solid"),
    "medium": PatternFill(start_color="FFDD57", end_color="FFDD57", fill_type="solid"),
    "low": PatternFill(start_color="85E085", end_color="85E085", fill_type="solid"),
}
SEV_FONTS = {
    "critical": Font(bold=True, color="FFFFFF", size=10),
    "high": Font(bold=True, color="FFFFFF", size=10),
    "medium": Font(bold=True, color="333333", size=10),
    "low": Font(bold=True, color="333333", size=10),
}

LOCATION_FILL = PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid")  # light purple


def auto_fit_columns(ws, total_cols, last_row, max_width=45, min_width=10):
    """Auto-fit column widths based on content."""
    for c_idx in range(1, total_cols + 1):
        header_val = ws.cell(row=1, column=c_idx).value
        max_len = len(str(header_val)) if header_val else min_width
        for r_idx in range(2, min(last_row + 1, 202)):
            val = ws.cell(row=r_idx, column=c_idx).value
            if val:
                max_len = max(max_len, min(len(str(val)), max_width))
        ws.column_dimensions[get_column_letter(c_idx)].width = max(min(max_len + 2, max_width), min_width)


# ---------------------------------------------------------------------------
# Tab 1: CVE Scan - Enriched
# ---------------------------------------------------------------------------
def build_cve_tab(wb, input_csv):
    """Build the CVE Scan tab from the enriched CSV."""
    with open(input_csv, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        header = next(reader)
        rows = list(reader)

    col_names = {c.strip(): i for i, c in enumerate(header)}
    file_col = col_names.get("Affected File/Path")
    func_col = col_names.get("Affected Function(s)")
    sub_col = col_names.get("Affected Subsystem")

    if any(c is None for c in (file_col, func_col, sub_col)):
        print("ERROR: enriched columns not found. Run extract_cve_details.py first.")
        sys.exit(1)

    # Reorder columns: move the 3 affected columns right after Severity (B)
    enriched_indices = [file_col, func_col, sub_col]
    sev_idx = col_names.get("Severity", 1)
    insert_after = sev_idx

    original_order = list(range(len(header)))
    remaining = [i for i in original_order if i not in enriched_indices]
    new_order = remaining[:insert_after + 1] + enriched_indices + remaining[insert_after + 1:]

    header = [header[i] for i in new_order]
    rows = [[row[i] if i < len(row) else "" for i in new_order] for row in rows]

    col_names = {c.strip(): i for i, c in enumerate(header)}
    file_col = col_names["Affected File/Path"]
    func_col = col_names["Affected Function(s)"]
    sub_col = col_names["Affected Subsystem"]
    file_col_1 = file_col + 1
    func_col_1 = func_col + 1
    sub_col_1 = sub_col + 1

    ws = wb.active
    ws.title = "CVE Scan - Enriched"
    total_cols = len(header)

    # Write header
    for c_idx, col_name in enumerate(header, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if c_idx == file_col_1:
            cell.fill = AZURE_HDR
        elif c_idx == func_col_1:
            cell.fill = GREEN_HDR
        elif c_idx == sub_col_1:
            cell.fill = ORANGE_HDR
        else:
            cell.fill = GREY_HDR

    # Write data rows
    sev_col_idx = col_names.get("Severity")
    for r_idx, row in enumerate(rows, start=2):
        for c_idx in range(1, total_cols + 1):
            value = row[c_idx - 1] if c_idx - 1 < len(row) else ""
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if c_idx == file_col_1:
                cell.fill = AZURE_FILL
            elif c_idx == func_col_1:
                cell.fill = GREEN_FILL
            elif c_idx == sub_col_1:
                cell.fill = ORANGE_FILL
            if sev_col_idx is not None and c_idx == sev_col_idx + 1:
                sev = value.strip().lower()
                if sev in SEV_FILLS:
                    cell.fill = SEV_FILLS[sev]
                    cell.font = SEV_FONTS[sev]

    last_row = len(rows) + 1
    ws.auto_filter.ref = f"A1:{get_column_letter(total_cols)}{last_row}"

    # Subsystem dropdown
    subsystems = sorted(
        {row[sub_col].strip() for row in rows if sub_col < len(row) and row[sub_col].strip()}
    )
    if subsystems:
        formula_list = ",".join(subsystems)
        if len(formula_list) <= 255:
            dv = DataValidation(type="list", formula1=f'"{formula_list}"', allow_blank=True, showDropDown=False)
        else:
            ws_helper = wb.create_sheet("_Subsystems")
            for i, s in enumerate(subsystems, start=1):
                ws_helper.cell(row=i, column=1, value=s)
            ws_helper.sheet_state = "hidden"
            dv = DataValidation(type="list", formula1=f"=_Subsystems!$A$1:$A${len(subsystems)}", allow_blank=True, showDropDown=False)
        dv.prompt = "Select a subsystem to filter"
        dv.promptTitle = "Subsystem Filter"
        sub_letter = get_column_letter(sub_col_1)
        dv.add(f"{sub_letter}2:{sub_letter}{last_row}")
        ws.add_data_validation(dv)

    # Severity dropdown
    if sev_col_idx is not None:
        severities = sorted(
            {row[sev_col_idx].strip() for row in rows if sev_col_idx < len(row) and row[sev_col_idx].strip()}
        )
        if severities:
            dv_sev = DataValidation(type="list", formula1=f'"{",".join(severities)}"', allow_blank=True, showDropDown=False)
            dv_sev.prompt = "Select a severity to filter"
            dv_sev.promptTitle = "Severity Filter"
            sev_letter = get_column_letter(sev_col_idx + 1)
            dv_sev.add(f"{sev_letter}2:{sev_letter}{last_row}")
            ws.add_data_validation(dv_sev)

    ws.freeze_panes = "A2"
    auto_fit_columns(ws, total_cols, last_row)

    return len(rows), len(subsystems)


# ---------------------------------------------------------------------------
# Tab 2: SAST Report
# ---------------------------------------------------------------------------
def build_sast_tab(wb, sast_file):
    """Build the SAST Report tab from a Finite State .numbers export."""
    from numbers_parser import Document

    doc = Document(sast_file)
    sheet = doc.sheets[0]
    table = sheet.tables[0]

    # Read all data
    sast_header = []
    for c in range(table.num_cols):
        val = table.cell(0, c).value
        # Clean up column names
        name = str(val).replace("columns.", "").replace("_", " ").title()
        sast_header.append(name)

    sast_rows = []
    for r in range(1, table.num_rows):
        row = []
        for c in range(table.num_cols):
            val = table.cell(r, c).value
            if val is None:
                row.append("")
            else:
                row.append(str(val))
        sast_rows.append(row)

    # Identify key columns
    sast_col = {name: i for i, name in enumerate(sast_header)}
    sev_col = sast_col.get("Severity")
    loc_col = sast_col.get("Location")

    ws = wb.create_sheet("SAST Report")
    total_cols = len(sast_header)

    # Write header
    for c_idx, col_name in enumerate(sast_header, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if loc_col is not None and c_idx == loc_col + 1:
            cell.fill = PURPLE_HDR
        elif sev_col is not None and c_idx == sev_col + 1:
            cell.fill = ORANGE_HDR
        else:
            cell.fill = GREY_HDR

    # Write data rows
    for r_idx, row in enumerate(sast_rows, start=2):
        for c_idx in range(1, total_cols + 1):
            value = row[c_idx - 1] if c_idx - 1 < len(row) else ""
            # Clean up empty-looking values
            if value in ("None", "[]"):
                value = ""
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)

            # Location column - light purple
            if loc_col is not None and c_idx == loc_col + 1:
                cell.fill = LOCATION_FILL

            # Severity colour
            if sev_col is not None and c_idx == sev_col + 1:
                sev = value.strip().lower()
                if sev in SEV_FILLS:
                    cell.fill = SEV_FILLS[sev]
                    cell.font = SEV_FONTS[sev]

    last_row = len(sast_rows) + 1
    ws.auto_filter.ref = f"A1:{get_column_letter(total_cols)}{last_row}"

    # Severity dropdown
    if sev_col is not None:
        severities = sorted(
            {row[sev_col].strip().lower() for row in sast_rows if sev_col < len(row) and row[sev_col].strip() and row[sev_col] != "None"}
        )
        if severities:
            dv_sev = DataValidation(type="list", formula1=f'"{",".join(severities)}"', allow_blank=True, showDropDown=False)
            dv_sev.prompt = "Select a severity"
            dv_sev.promptTitle = "Severity Filter"
            sev_letter = get_column_letter(sev_col + 1)
            dv_sev.add(f"{sev_letter}2:{sev_letter}{last_row}")
            ws.add_data_validation(dv_sev)

    ws.freeze_panes = "A2"
    auto_fit_columns(ws, total_cols, last_row)

    return len(sast_rows)


# ---------------------------------------------------------------------------
# GUI file browser
# ---------------------------------------------------------------------------
def browse_files():
    """Simple tkinter GUI to select the CVE CSV and optional SAST .numbers file."""
    import tkinter as tk
    from tkinter import filedialog, messagebox

    result = {"csv": None, "sast": None}

    root = tk.Tk()
    root.title("Finite State Report Builder")
    root.geometry("620x280")
    root.resizable(False, False)

    # --- CVE CSV ---
    tk.Label(root, text="Enriched CVE CSV (required):", font=("Segoe UI", 10, "bold")).place(x=20, y=20)
    csv_var = tk.StringVar()
    csv_entry = tk.Entry(root, textvariable=csv_var, width=55, font=("Segoe UI", 9))
    csv_entry.place(x=20, y=50)

    def browse_csv():
        path = filedialog.askopenfilename(
            title="Select Enriched CVE CSV",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        )
        if path:
            csv_var.set(path)

    tk.Button(root, text="Browse...", command=browse_csv, width=10).place(x=510, y=47)

    # --- SAST .numbers ---
    tk.Label(root, text="SAST Findings .numbers file (optional):", font=("Segoe UI", 10, "bold")).place(x=20, y=100)
    sast_var = tk.StringVar()
    sast_entry = tk.Entry(root, textvariable=sast_var, width=55, font=("Segoe UI", 9))
    sast_entry.place(x=20, y=130)

    def browse_sast():
        path = filedialog.askopenfilename(
            title="Select SAST .numbers File",
            filetypes=[("Numbers files", "*.numbers"), ("All files", "*.*")],
        )
        if path:
            sast_var.set(path)

    tk.Button(root, text="Browse...", command=browse_sast, width=10).place(x=510, y=127)

    # --- Generate button ---
    def on_generate():
        csv_path = csv_var.get().strip()
        if not csv_path:
            messagebox.showerror("Error", "Please select the enriched CVE CSV file.")
            return
        result["csv"] = csv_path
        sast_path = sast_var.get().strip()
        if sast_path:
            result["sast"] = sast_path
        root.destroy()

    tk.Button(root, text="Generate Excel Report", command=on_generate,
              font=("Segoe UI", 11, "bold"), bg="#4472C4", fg="white",
              width=25, height=2).place(x=180, y=190)

    def on_close():
        result["csv"] = None
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_close)
    root.mainloop()

    return result["csv"], result["sast"]


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    csv_file = None
    sast_file = None

    # Check for command-line args or launch GUI
    if len(sys.argv) >= 2:
        csv_file = sys.argv[1]
        if len(sys.argv) >= 3:
            sast_file = sys.argv[2]
    else:
        # Launch GUI browser
        csv_file, sast_file = browse_files()

    if not csv_file:
        print("No CSV file selected. Exiting.")
        sys.exit(0)

    output_xlsx = str(Path(csv_file).with_suffix(".xlsx"))

    wb = Workbook()

    # Tab 1: CVE Scan
    print("Building CVE Scan tab...")
    cve_rows, num_subsystems = build_cve_tab(wb, csv_file)
    print(f"  {cve_rows} CVE rows, {num_subsystems} unique subsystems")

    # Tab 2: SAST Report (optional)
    sast_rows = 0
    if sast_file:
        print(f"Building SAST Report tab from: {sast_file}")
        sast_rows = build_sast_tab(wb, sast_file)
        print(f"  {sast_rows} SAST findings")

    wb.save(output_xlsx)
    print(f"\nExcel report saved to: {output_xlsx}")
    print(f"  Tab 1: CVE Scan - Enriched ({cve_rows} rows)")
    if sast_rows:
        print(f"  Tab 2: SAST Report ({sast_rows} rows)")


if __name__ == "__main__":
    main()
