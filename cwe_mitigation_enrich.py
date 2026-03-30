"""
Phase 5 - CWE Mitigation Enrichment (air-gap compatible)

Adds a "CWE Mitigation" column to scan_for_Elta_enriched.csv by:
  1. Looking up each row's CWE ID in the local CWE XML database
  2. Walking up the CWE parent chain when a CWE has no own mitigations
  3. For rows missing a CWE entirely (CWE-0 / CWE--1), resolving the
     real CWE from local NVD JSON 2.0 feeds, then doing the lookup

Required local databases (in cwe_db/):
  - CWE XML:     cwec_latest.xml.zip  (from cwe.mitre.org)
  - NVD feeds:   nvd_feeds/nvdcve-2.0-YYYY.json.gz  (from nvd.nist.gov)

On a connected system, use --download to fetch everything automatically.
On an air-gapped system, transfer the files from a sanitisation server.

Usage:
    python cwe_mitigation_enrich.py                   (uses local DBs)
    python cwe_mitigation_enrich.py --download        (downloads DBs first)

Updates scan_for_Elta_enriched.csv in-place (backs up original first)
and regenerates the Excel report.
"""

import csv
import glob
import gzip
import json
import os
import shutil
import sys
import urllib.request
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
PROJECT_DIR = Path(r"C:\Amp_demos\Elta-Finitestate")
INPUT_CSV = PROJECT_DIR / "scan_for_Elta_enriched.csv"
CWE_DB_DIR = PROJECT_DIR / "cwe_db"
NVD_FEEDS_DIR = CWE_DB_DIR / "nvd_feeds"
CWE_ZIP_URL = "https://cwe.mitre.org/data/xml/cwec_latest.xml.zip"
NVD_FEED_URL = "https://nvd.nist.gov/feeds/json/cve/2.0/nvdcve-2.0-{year}.json.gz"
CWE_NS = "{http://cwe.mitre.org/cwe-7}"

MITIGATION_COL = "CWE Mitigation"
MAX_PARENT_DEPTH = 5  # max levels to walk up the CWE hierarchy

# ---------------------------------------------------------------------------
# Fallback mitigations: for CWEs with no mitigations in the DB and for CVEs
# where NVD has no CWE at all (NVD-CWE-noinfo).  Derived from description
# keyword analysis and authoritative CWE guidance.
# ---------------------------------------------------------------------------
# CWEs that exist in the DB but have no mitigation text anywhere in their chain
MANUAL_CWE_MITIGATIONS = {
    "835": "[Implementation] Ensure all loops have reachable exit conditions. Add iteration limits, timeouts, or watchdog mechanisms to prevent infinite loops. Validate loop termination conditions during code review.",
    "755": "[Implementation] Implement proper exception and error handling for all exceptional conditions. Use structured error handling (try/catch) and ensure errors are logged and handled gracefully without exposing internal state.",
    "843": "[Implementation] Enforce strict type checking on all inputs and internal data. Validate object types before casting or dereferencing. Use type-safe languages or compiler options where possible.",
    "834": "[Implementation] Add upper bounds or iteration limits to all loops processing external data. Implement timeouts for long-running operations to prevent resource exhaustion.",
    "388": "[Implementation] Handle all error conditions explicitly. Do not suppress or ignore exceptions. Ensure error-handling code does not introduce new vulnerabilities.",
    "911": "[Implementation] Use proper synchronization primitives (mutexes, locks, atomics) when accessing shared data. Ensure lock ordering is consistent to prevent deadlocks.",
    "668": "[Architecture and Design] Apply the principle of least privilege. Restrict resource access to only the components that require it. Use access control lists and capability-based security.",
    "704": "[Implementation] Validate types before conversion or casting. Use safe casting functions that check for compatibility. Avoid implicit type conversions on untrusted data.",
    "693": "[Architecture and Design] Do not rely solely on a single protection mechanism. Implement defense-in-depth with multiple complementary security layers.",
    "670": "[Implementation] Ensure the implementation follows the documented API contract. Validate return values and state transitions match expected behavior.",
    "345": "[Architecture and Design] Verify the authenticity and integrity of all data and messages before processing. Use cryptographic signatures or MACs to validate origin.",
    "229": "[Implementation] Validate all input values, including checking for missing or incomplete values. Do not assume optional fields will be present.",
    "189": "[Implementation] Use safe integer arithmetic. Check for overflow/underflow before performing arithmetic operations. Use fixed-width integer types with known bounds.",
    "706": "[Architecture and Design] Use canonical forms when comparing or resolving resource identifiers. Normalize paths, names, and references before use.",
    "436": "[Architecture and Design] Resolve conflicting interpretations by enforcing a single canonical interpretation. Reject ambiguous inputs.",
    "399": "[Implementation] Implement proper resource lifecycle management. Ensure all allocated resources (memory, file handles, connections) are released on all code paths, including error paths.",
    "703": "[Implementation] Implement comprehensive error and exception handling. Ensure all failure modes are addressed and the system fails safely.",
    "697": "[Implementation] Use exact matching rather than loose comparison when validating security-critical values. Avoid type coercion in comparisons.",
    "16":  "[Architecture and Design] Ensure secure default configurations. Apply the principle of least privilege to all configurable settings. Validate configuration values against an allowlist.",
    "347": "[Implementation] Verify digital signatures or MACs on all received data before processing. Reject data that fails integrity verification.",
}

# Keyword-based fallback for CVEs with no CWE at all (NVD-CWE-noinfo)
DESCRIPTION_FALLBACKS = [
    ("use-after-free", "[Implementation] Ensure proper object lifecycle management. Nullify pointers after freeing. Use smart pointers or ownership models to prevent use-after-free. Applies to: use-after-free vulnerability."),
    ("null pointer dereference", "[Implementation] Validate all pointers before dereferencing. Check return values of functions that may return NULL. Enable compiler warnings for potential null dereferences. Applies to: null pointer dereference."),
    ("null pointer", "[Implementation] Validate all pointers before dereferencing. Check return values of functions that may return NULL. Enable compiler warnings for potential null dereferences. Applies to: null pointer issue."),
    ("buffer overflow", "[Implementation] Use bounds-checked functions (e.g., strncpy, snprintf). Validate buffer sizes before copy operations. Enable stack canaries and ASLR. Applies to: buffer overflow."),
    ("heap overflow", "[Implementation] Validate allocation sizes and copy lengths. Use bounds-checked memory operations. Enable heap protection mechanisms (ASLR, guard pages). Applies to: heap overflow."),
    ("stack overflow", "[Implementation] Limit recursion depth. Validate stack-consuming inputs. Use iterative algorithms where possible. Applies to: stack overflow."),
    ("out-of-bounds read", "[Implementation] Validate array indices and buffer offsets before access. Use bounds-checked data structures. Enable address sanitizers during testing. Applies to: out-of-bounds read."),
    ("out-of-bounds write", "[Implementation] Validate indices and lengths before write operations. Use bounds-checked functions. Enable memory sanitizers during testing. Applies to: out-of-bounds write."),
    ("out-of-bounds", "[Implementation] Validate all array indices and buffer boundaries before access. Use bounds-checked APIs and enable sanitizers during testing. Applies to: out-of-bounds access."),
    ("integer overflow", "[Implementation] Use safe integer arithmetic with overflow checks. Validate input ranges before arithmetic operations. Use compiler built-ins for overflow detection. Applies to: integer overflow."),
    ("integer underflow", "[Implementation] Validate that subtraction operands will not produce negative values in unsigned arithmetic. Use safe arithmetic functions with underflow checks. Applies to: integer underflow."),
    ("infinite loop", "[Implementation] Ensure all loops have reachable exit conditions. Add iteration limits and timeouts. Validate loop control variables against external input. Applies to: infinite loop / excessive iteration."),
    ("denial of service", "[Architecture and Design] Implement rate limiting, resource quotas, and input size validation. Add timeouts for all blocking operations. Isolate critical services from untrusted input paths. Applies to: denial of service."),
    ("race condition", "[Implementation] Use proper synchronization primitives (mutexes, locks, atomics) when accessing shared resources. Follow consistent lock ordering. Applies to: race condition."),
    ("deadlock", "[Implementation] Ensure consistent lock acquisition ordering across all code paths. Use lock timeouts and deadlock detection mechanisms. Applies to: deadlock."),
    ("double free", "[Implementation] Set pointers to NULL after freeing. Use ownership models to ensure each allocation has a single owner responsible for deallocation. Applies to: double free."),
    ("memory leak", "[Implementation] Ensure all allocated memory is freed on every code path including error paths. Use RAII patterns or smart pointers. Run leak detection tools regularly. Applies to: memory leak."),
    ("memory corruption", "[Implementation] Use bounds-checked memory operations. Enable memory protection features (ASLR, stack canaries, guard pages). Validate all indices and sizes. Applies to: memory corruption."),
    ("uninitialized", "[Implementation] Initialize all variables before use. Enable compiler warnings for uninitialized variables. Use static analysis tools to detect uninitialized reads. Applies to: uninitialized memory/variable."),
    ("information disclosure", "[Architecture and Design] Apply the principle of least privilege to data access. Clear sensitive data from memory after use. Validate that responses do not leak internal state. Applies to: information disclosure."),
    ("information leak", "[Architecture and Design] Apply the principle of least privilege to data access. Clear sensitive data from memory after use. Validate that responses do not leak internal state. Applies to: information leak."),
    ("privilege escalation", "[Architecture and Design] Enforce proper access control checks on all privileged operations. Apply the principle of least privilege. Validate user permissions before granting elevated access. Applies to: privilege escalation."),
    ("type confusion", "[Implementation] Enforce strict type checking on all inputs and internal data. Validate object types before casting. Use type-safe languages or compiler options. Applies to: type confusion."),
    ("overflow", "[Implementation] Validate all buffer sizes, array indices, and arithmetic operands. Use bounds-checked functions and safe integer arithmetic. Applies to: overflow."),
    ("underflow", "[Implementation] Validate arithmetic operands to ensure results remain within expected ranges. Use safe arithmetic functions. Applies to: underflow."),
    ("crash", "[Implementation] Add defensive null checks, bounds validation, and error handling on failure paths to prevent crashes. Apply the relevant kernel patch. Applies to: crash."),
    ("hang", "[Implementation] Add timeouts and watchdog mechanisms to prevent system hangs. Validate loop termination conditions and lock ordering. Applies to: system hang."),
    ("panic", "[Implementation] Add defensive checks to prevent kernel panics. Validate inputs and state before operations that may trigger panic. Apply the relevant kernel patch. Applies to: kernel panic."),
    ("assertion", "[Implementation] Validate preconditions before triggering assertions. Ensure assertion conditions cannot be influenced by untrusted input. Applies to: assertion failure."),
    ("warning", "[Implementation] Address root cause of kernel warnings. Validate state and inputs to prevent unexpected conditions. Apply the relevant kernel patch. Applies to: kernel warning."),
    ("oops", "[Implementation] Fix the root cause of the kernel oops. Add null checks and bounds validation. Apply the relevant kernel patch. Applies to: kernel oops."),
]

# Generic fallback for Linux kernel CVEs with no specific keyword match
GENERIC_KERNEL_FALLBACK = "[Implementation] Apply the vendor-provided kernel patch to the affected component. Update the Linux kernel to a version containing the fix. If patching is not immediately possible, evaluate whether the affected subsystem can be disabled or access-restricted as a temporary mitigation."


# ---------------------------------------------------------------------------
# Download helpers
# ---------------------------------------------------------------------------
def download_cwe_db():
    """Download and extract the CWE XML database."""
    os.makedirs(CWE_DB_DIR, exist_ok=True)
    zip_path = CWE_DB_DIR / "cwec_latest.xml.zip"
    print(f"Downloading CWE database from {CWE_ZIP_URL} ...")
    urllib.request.urlretrieve(CWE_ZIP_URL, str(zip_path))
    print(f"  Downloaded: {zip_path} ({zip_path.stat().st_size:,} bytes)")
    with zipfile.ZipFile(zip_path, "r") as z:
        z.extractall(CWE_DB_DIR)
    print(f"  Extracted to: {CWE_DB_DIR}")


def download_nvd_feeds(years):
    """Download NVD JSON 2.0 feeds for the given years."""
    os.makedirs(NVD_FEEDS_DIR, exist_ok=True)
    for year in years:
        gz_path = NVD_FEEDS_DIR / f"nvdcve-2.0-{year}.json.gz"
        if gz_path.exists():
            print(f"  NVD {year}: already downloaded")
            continue
        url = NVD_FEED_URL.format(year=year)
        print(f"  NVD {year}: downloading ... ", end="", flush=True)
        try:
            urllib.request.urlretrieve(url, str(gz_path))
            size_mb = gz_path.stat().st_size / 1024 / 1024
            print(f"{size_mb:.1f} MB")
        except Exception as e:
            print(f"FAILED: {e}")


# ---------------------------------------------------------------------------
# CWE Database loader (with parent chain)
# ---------------------------------------------------------------------------
def find_cwe_xml():
    """Locate the CWE XML file in the cwe_db directory."""
    pattern = str(CWE_DB_DIR / "cwec_v*.xml")
    matches = glob.glob(pattern)
    if matches:
        return max(matches)
    pattern = str(CWE_DB_DIR / "*.xml")
    matches = [m for m in glob.glob(pattern) if not m.endswith(".zip")]
    if matches:
        return matches[0]
    return None


def load_cwe_database(xml_path):
    """Parse the CWE XML and return mitigations dict and parent-chain map."""
    print(f"Loading CWE database: {xml_path}")
    tree = ET.parse(xml_path)
    root = tree.getroot()

    mitigations = {}   # CWE ID -> mitigation text
    parents = {}       # CWE ID -> list of parent CWE IDs
    names = {}         # CWE ID -> name

    for weakness in root.iter(CWE_NS + "Weakness"):
        cwe_id = weakness.get("ID")
        names[cwe_id] = weakness.get("Name", "")

        # Extract parent relationships
        rels_el = weakness.find(CWE_NS + "Related_Weaknesses")
        if rels_el is not None:
            parent_ids = []
            for rel in rels_el.findall(CWE_NS + "Related_Weakness"):
                if rel.get("Nature") == "ChildOf":
                    pid = rel.get("CWE_ID")
                    if pid and pid not in parent_ids:
                        parent_ids.append(pid)
            if parent_ids:
                parents[cwe_id] = parent_ids

        # Extract mitigations
        mits_el = weakness.find(CWE_NS + "Potential_Mitigations")
        if mits_el is None:
            continue

        mit_texts = []
        for mit in mits_el.findall(CWE_NS + "Mitigation"):
            phase_el = mit.find(CWE_NS + "Phase")
            phase = phase_el.text.strip() if phase_el is not None else ""
            desc_el = mit.find(CWE_NS + "Description")
            if desc_el is None:
                continue
            desc = ET.tostring(desc_el, encoding="unicode", method="text").strip()
            if not desc:
                continue
            if phase:
                mit_texts.append(f"[{phase}] {desc}")
            else:
                mit_texts.append(desc)

        if mit_texts:
            mitigations[cwe_id] = " | ".join(mit_texts)

    print(f"  Loaded {len(mitigations)} CWEs with own mitigations")
    print(f"  Loaded {len(parents)} CWEs with parent relationships")
    return mitigations, parents, names


def resolve_mitigation(cwe_num, mitigations, parents):
    """Look up mitigation for a CWE, walking up the parent chain if needed."""
    if cwe_num in mitigations:
        return mitigations[cwe_num], cwe_num, False

    # Walk up parent chain (BFS)
    visited = {cwe_num}
    queue = [cwe_num]
    for _ in range(MAX_PARENT_DEPTH):
        next_queue = []
        for cid in queue:
            for pid in parents.get(cid, []):
                if pid in visited:
                    continue
                visited.add(pid)
                if pid in mitigations:
                    return mitigations[pid], pid, True
                next_queue.append(pid)
        if not next_queue:
            break
        queue = next_queue

    return "", "", False


# ---------------------------------------------------------------------------
# NVD feed loader (CVE -> CWE mapping)
# ---------------------------------------------------------------------------
def load_nvd_cwe_map(cve_ids_needed):
    """Load CWE mappings from local NVD JSON feeds for the given CVE IDs."""
    # Determine which years we need
    years_needed = set()
    for cve_id in cve_ids_needed:
        parts = cve_id.split("-")
        if len(parts) >= 2:
            years_needed.add(parts[1])

    cve_to_cwe = {}
    feeds_found = 0

    for year in sorted(years_needed):
        gz_path = NVD_FEEDS_DIR / f"nvdcve-2.0-{year}.json.gz"
        if not gz_path.exists():
            print(f"  NVD feed {year}: not found, skipping")
            continue
        feeds_found += 1
        print(f"  NVD feed {year}: loading ... ", end="", flush=True)
        with gzip.open(gz_path, "rt", encoding="utf-8") as f:
            data = json.load(f)

        matched = 0
        for vuln in data.get("vulnerabilities", []):
            cve_data = vuln.get("cve", {})
            cve_id = cve_data.get("id", "")
            if cve_id not in cve_ids_needed:
                continue

            # Extract CWE from weaknesses
            for w in cve_data.get("weaknesses", []):
                for desc in w.get("description", []):
                    val = desc.get("value", "")
                    if val.startswith("CWE-") and val not in ("CWE-noinfo", "CWE-Other"):
                        cve_to_cwe[cve_id] = val
                        matched += 1
                        break
                if cve_id in cve_to_cwe:
                    break

        print(f"{matched} CWEs resolved")

    if feeds_found == 0:
        print("  No NVD feeds found in cwe_db/nvd_feeds/")
    return cve_to_cwe


# ---------------------------------------------------------------------------
# CSV enrichment
# ---------------------------------------------------------------------------
def enrich_csv(mitigations, parents, names):
    """Add the CWE Mitigation column to the enriched CSV."""
    with open(INPUT_CSV, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        header = next(reader)
        rows = list(reader)

    col = {c.strip(): i for i, c in enumerate(header)}

    if "CWE" not in col:
        print("ERROR: 'CWE' column not found in CSV.")
        sys.exit(1)

    cwe_idx = col["CWE"]
    cve_idx = col["CVE ID"]

    # Add or update the mitigation column
    if MITIGATION_COL in col:
        mit_idx = col[MITIGATION_COL]
        print(f"  Updating existing '{MITIGATION_COL}' column (index {mit_idx})")
    else:
        mit_idx = len(header)
        header.append(MITIGATION_COL)
        for row in rows:
            row.append("")
        print(f"  Adding new '{MITIGATION_COL}' column (index {mit_idx})")

    # --- Step 1: Identify rows missing CWE and resolve from NVD feeds ---
    no_cwe_cves = set()
    for row in rows:
        cwe_raw = row[cwe_idx].strip() if cwe_idx < len(row) else ""
        if not cwe_raw or cwe_raw in ("CWE-0", "CWE--1"):
            cve_id = row[cve_idx].strip()
            if cve_id:
                no_cwe_cves.add(cve_id)

    nvd_cwe_map = {}
    if no_cwe_cves:
        print(f"\n  Resolving CWE for {len(no_cwe_cves)} CVEs missing CWE ID ...")
        nvd_cwe_map = load_nvd_cwe_map(no_cwe_cves)
        print(f"  Resolved {len(nvd_cwe_map)} / {len(no_cwe_cves)} from NVD feeds")

    # --- Step 2: Enrich all rows ---
    desc_idx = col.get("Description")
    stats = {
        "total": 0, "direct": 0, "parent": 0, "nvd_resolved": 0,
        "manual_cwe": 0, "desc_fallback": 0, "no_mitigation": 0,
    }

    for row in rows:
        stats["total"] += 1
        cwe_raw = row[cwe_idx].strip() if cwe_idx < len(row) else ""
        cve_id = row[cve_idx].strip()
        description = row[desc_idx].strip() if desc_idx is not None and desc_idx < len(row) else ""
        had_no_cwe = False

        # Try NVD resolution for missing CWEs
        if not cwe_raw or cwe_raw in ("CWE-0", "CWE--1"):
            if cve_id in nvd_cwe_map:
                cwe_raw = nvd_cwe_map[cve_id]
                row[cwe_idx] = cwe_raw
            else:
                had_no_cwe = True

        # Try CWE-based lookup
        mit_text = ""
        source_label = ""

        if cwe_raw and cwe_raw not in ("CWE-0", "CWE--1"):
            cwe_num = cwe_raw.replace("CWE-", "").strip()

            # 1) CWE DB (direct or parent chain)
            db_mit, source_cwe, from_parent = resolve_mitigation(
                cwe_num, mitigations, parents
            )
            if db_mit:
                if from_parent:
                    source_name = names.get(source_cwe, "")
                    mit_text = f"[via CWE-{source_cwe}: {source_name}] {db_mit}"
                else:
                    mit_text = db_mit

                if cve_id in nvd_cwe_map:
                    stats["nvd_resolved"] += 1
                elif from_parent:
                    stats["parent"] += 1
                else:
                    stats["direct"] += 1

            # 2) Manual CWE mitigations (hand-authored for known gaps)
            elif cwe_num in MANUAL_CWE_MITIGATIONS:
                mit_text = MANUAL_CWE_MITIGATIONS[cwe_num]
                stats["manual_cwe"] += 1

        # 3) Description-based fallback
        if not mit_text and description:
            desc_lower = description.lower()
            for keyword, fallback_mit in DESCRIPTION_FALLBACKS:
                if keyword in desc_lower:
                    mit_text = fallback_mit
                    stats["desc_fallback"] += 1
                    break

        # 4) Generic fallback (last resort)
        if not mit_text and description:
            desc_lower = description.lower()
            if "rejected" in desc_lower and "withdrawn" in desc_lower:
                mit_text = "This CVE has been rejected/withdrawn by its CVE Numbering Authority. No mitigation required."
                stats["desc_fallback"] += 1
            elif "linux kernel" in desc_lower or "in the kernel" in desc_lower:
                mit_text = GENERIC_KERNEL_FALLBACK
                stats["desc_fallback"] += 1
            elif "802.11" in desc_lower or "wpa" in desc_lower or "wi-fi" in desc_lower or "wifi" in desc_lower:
                mit_text = "[Architecture and Design] Update Wi-Fi firmware and drivers to versions that address the vulnerability. Enforce WPA3 where possible. Apply vendor-provided patches for affected wireless stack components."
                stats["desc_fallback"] += 1
            else:
                mit_text = "[Implementation] Apply the vendor-provided patch or update the affected component to a fixed version. If patching is not immediately possible, evaluate compensating controls such as restricting access to the affected functionality."
                stats["desc_fallback"] += 1

        if mit_text:
            row[mit_idx] = mit_text
        else:
            stats["no_mitigation"] += 1

    # Backup and save
    backup = str(INPUT_CSV) + ".bak_mit"
    shutil.copy2(INPUT_CSV, backup)
    print(f"  Backup saved: {backup}")

    with open(INPUT_CSV, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(rows)

    total_with = (stats["direct"] + stats["parent"] + stats["nvd_resolved"]
                  + stats["manual_cwe"] + stats["desc_fallback"])
    print(f"\n{'='*60}")
    print(f"  Total CVE rows:              {stats['total']}")
    print(f"  Direct CWE match:            {stats['direct']}")
    print(f"  Via parent CWE chain:        {stats['parent']}")
    print(f"  Resolved CWE from NVD feed:  {stats['nvd_resolved']}")
    print(f"  Manual CWE mitigations:      {stats['manual_cwe']}")
    print(f"  Description-based fallback:  {stats['desc_fallback']}")
    print(f"  -----------------------------------")
    print(f"  Total with mitigation:       {total_with}")
    print(f"  No mitigation available:     {stats['no_mitigation']}")
    print(f"  Coverage:                    {total_with*100//stats['total']}%")
    print(f"  Output: {INPUT_CSV}")


# ---------------------------------------------------------------------------
# Excel regeneration (reuses create_excel_report.py logic)
# ---------------------------------------------------------------------------
def rebuild_excel():
    """Regenerate the Excel report to include the new column."""
    xlsx_path = INPUT_CSV.with_suffix(".xlsx")
    print(f"\nRegenerating Excel report: {xlsx_path}")

    sys.path.insert(0, str(PROJECT_DIR))
    from create_excel_report import build_cve_tab

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    cve_rows, num_subsystems = build_cve_tab(wb, str(INPUT_CSV))

    # Style the Mitigation column
    ws = wb.active
    header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    if MITIGATION_COL in header_row:
        mit_col_1 = header_row.index(MITIGATION_COL) + 1
        TEAL_HDR = PatternFill(start_color="008080", end_color="008080", fill_type="solid")
        TEAL_FILL = PatternFill(start_color="E0F2F1", end_color="E0F2F1", fill_type="solid")
        ws.cell(row=1, column=mit_col_1).fill = TEAL_HDR

        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=mit_col_1)
            cell.fill = TEAL_FILL
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        ws.column_dimensions[get_column_letter(mit_col_1)].width = 60

    wb.save(str(xlsx_path))
    print(f"  Excel report saved: {xlsx_path}")
    print(f"  {cve_rows} CVE rows, {num_subsystems} unique subsystems")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    # Determine which NVD feed years are needed
    needed_years = set()
    with open(INPUT_CSV, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        header = next(reader)
        col = {c.strip(): i for i, c in enumerate(header)}
        cwe_idx = col["CWE"]
        cve_idx = col["CVE ID"]
        for row in reader:
            cwe = row[cwe_idx].strip() if cwe_idx < len(row) else ""
            if not cwe or cwe in ("CWE-0", "CWE--1"):
                parts = row[cve_idx].split("-")
                if len(parts) >= 2:
                    needed_years.add(parts[1])

    if "--download" in sys.argv:
        download_cwe_db()
        if needed_years:
            print(f"\nDownloading NVD feeds for years: {sorted(needed_years)}")
            download_nvd_feeds(sorted(needed_years))

    xml_path = find_cwe_xml()
    if not xml_path:
        print("ERROR: CWE XML database not found in cwe_db/")
        print("  Either run with --download, or place cwec_latest.xml.zip")
        print(f"  in: {CWE_DB_DIR}")
        sys.exit(1)

    mitigations, parents, names = load_cwe_database(xml_path)
    enrich_csv(mitigations, parents, names)
    rebuild_excel()


if __name__ == "__main__":
    main()
